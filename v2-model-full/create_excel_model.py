#!/usr/bin/env python3
"""Ultra-simple Excel model generator for survival credit modeling"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Create sample historical data
def create_sample_data():
    """Generate minimal sample data for 3 vintages"""
    data = []
    
    # 3 vintages, 24 months each
    for vintage in ['2023-01', '2023-02', '2023-03']:
        starting_bal = 1_000_000
        
        for month in range(1, 25):
            # Simple declining rates
            pay_rate = 0.02 * (1 - month * 0.002)  # Starts at 2%, declines
            co_rate = 0.005 * (1 + month * 0.001)   # Starts at 0.5%, increases
            
            payment = starting_bal * pay_rate
            chargeoff = starting_bal * co_rate
            ending_bal = starting_bal - payment - chargeoff
            
            data.append({
                'Vintage': vintage,
                'Month_Age': month,
                'Beginning_Balance': starting_bal,
                'Payment_Amt': payment,
                'Chargeoff_Amt': chargeoff,
                'Ending_Balance': ending_bal,
                'Loan_Count': 100,
                'Is_Actual': 1
            })
            
            starting_bal = ending_bal
    
    return pd.DataFrame(data)

# Create the Excel model
def create_excel_model():
    wb = Workbook()
    
    # 1. Raw_Data sheet
    ws_raw = wb.active
    ws_raw.title = "Raw_Data"
    df_raw = create_sample_data()
    
    for r in dataframe_to_rows(df_raw, index=False, header=True):
        ws_raw.append(r)
    
    # 2. Rate_Analysis sheet
    ws_rates = wb.create_sheet("Rate_Analysis")
    ws_rates.append(['Month_Age', 'Vintage_Count', 'Payment_Rate', 'CO_Rate'])
    
    for month in range(1, 25):
        row = month + 1
        ws_rates.append([
            month,
            f'=COUNTIFS(Raw_Data!$B:$B,A{row},Raw_Data!$H:$H,1)',
            f'=IFERROR(SUMIFS(Raw_Data!$D:$D,Raw_Data!$B:$B,A{row},Raw_Data!$H:$H,1)/SUMIFS(Raw_Data!$C:$C,Raw_Data!$B:$B,A{row},Raw_Data!$H:$H,1),0)',
            f'=IFERROR(SUMIFS(Raw_Data!$E:$E,Raw_Data!$B:$B,A{row},Raw_Data!$H:$H,1)/SUMIFS(Raw_Data!$C:$C,Raw_Data!$B:$B,A{row},Raw_Data!$H:$H,1),0)'
        ])
    
    # 3. Forecast_Rates sheet (extends to month 144)
    ws_forecast = wb.create_sheet("Forecast_Rates")
    ws_forecast.append(['Month_Age', 'Payment_Rate', 'CO_Rate', 'Source'])
    
    # First 24 months from historical
    for month in range(1, 25):
        row = month + 1
        ws_forecast.append([
            month,
            f'=Rate_Analysis!C{row}',
            f'=Rate_Analysis!D{row}',
            'Historical'
        ])
    
    # Months 25-144 with decay
    for month in range(25, 145):
        row = month + 1
        decay_factor = (month - 24) / 12
        ws_forecast.append([
            month,
            f'=B25*0.95^{decay_factor:.2f}',
            f'=C25*0.90^{decay_factor:.2f}',
            'Extended'
        ])
    
    # 4. Forecast_Output sheet (for new vintage 2025-01)
    ws_output = wb.create_sheet("Forecast_Output")
    ws_output.append(['Vintage', 'Month_Age', 'Beginning_Bal', 'Payment_Amt', 'CO_Amt', 'Ending_Bal', 'Payment_Rate', 'CO_Rate'])
    
    # Starting balance
    starting_balance = 1_000_000
    
    # First row
    ws_output.append([
        '2025-01',
        1,
        starting_balance,
        f'=C2*Forecast_Rates!B2',
        f'=C2*Forecast_Rates!C2',
        f'=C2-D2-E2',
        f'=Forecast_Rates!B2',
        f'=Forecast_Rates!C2'
    ])
    
    # Remaining rows
    for month in range(2, 145):
        row = month + 1
        ws_output.append([
            '2025-01',
            month,
            f'=F{row-1}',  # Beginning = Previous Ending
            f'=C{row}*Forecast_Rates!B{month+1}',
            f'=C{row}*Forecast_Rates!C{month+1}',
            f'=C{row}-D{row}-E{row}',
            f'=Forecast_Rates!B{month+1}',
            f'=Forecast_Rates!C{month+1}'
        ])
    
    # Save the file
    wb.save('/mnt/c/Users/clays/OneDrive/Documents/Github/survival-credit-modeling/v2-model-full/survival_model_v2.xlsx')
    print("Excel model created: survival_model_v2.xlsx")
    print("\nModel structure:")
    print("- Raw_Data: Sample historical data (3 vintages x 24 months)")
    print("- Rate_Analysis: Historical rate calculations")
    print("- Forecast_Rates: Rates extended to 144 months")
    print("- Forecast_Output: Single vintage forecast (2025-01)")

if __name__ == "__main__":
    create_excel_model()